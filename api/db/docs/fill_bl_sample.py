#!/usr/bin/env python3
# 填 Damon 原模版 (bl-sample-template.xlsx)，openpyxl 只改指定单元格，其余字节级保留（不重造）。
# 用法: python3 fill_bl_sample.py <template.xlsx>   数据从 stdin 读 JSON，xlsx 写 stdout。
import sys, json, io
from copy import copy
import openpyxl

tpl = sys.argv[1]
d = json.load(sys.stdin)
wb = openpyxl.load_workbook(tpl)
ws = wb.active

def n(v, dp=0):
    if v is None or v == "":
        return ""
    try:
        return f"{float(v):,.{dp}f}"
    except Exception:
        return str(v)

def s(addr, val):
    ws[addr] = "" if val is None else val

s("A2", (d.get("shipperName", "") or "") + (("\nADD: " + d["shipperAddrEn"]) if d.get("shipperAddrEn") else ""))
s("F2", d.get("blNo", ""))
s("F3", d.get("releaseType", "") or "SWB 海运单")
s("E4", "")  # 付款方式去掉（Damon 0813）
s("F4", "")
s("F5", d.get("hsCode", ""))
s("F6", "是 ( V )    否 (   )" if d.get("showHs") else "是 (   )    否 ( V )")
s("A6", (d.get("consignee", "") or "") + (("\n" + d["consAddr"]) if d.get("consAddr") else ""))
s("A10", d.get("notify", "") or "SAME AS CONSIGNEE")
s("A12", " ".join([x for x in [d.get("vessel", ""), d.get("voyage", "")] if x]))
s("E12", d.get("pol", ""))
s("A14", d.get("pod", ""))
s("E14", d.get("finalDest", "") or d.get("pod", ""))
s("A16", d.get("marks", "") or "N/M")
s("B16", (n(d["totalCtn"]) + " CARTONS") if d.get("totalCtn") else "")
s("C16", (d.get("description", "") or "") + (("\nHS: " + d["hsCode"]) if d.get("hsCode") else ""))
s("E16", (n(d["gwKg"], 2) + " KGS") if d.get("gwKg") else "")
s("G16", (n(d["cbm"], 3) + " CBM") if d.get("cbm") else "")
s("A21", "✓ 客户已确认提单信息（含 HS 显示选择）" if d.get("confirmed") else "⚠ 待客户确认提单信息（HS/货描）——确认前请勿提交")  # 双方确认状态
s("A22", "⚠ 付款方式 P/C 请确认后再发（成交方式需与客户核对）")
s("A23", "VGM 称重方式：Method 2 累加计算法（货重 = 净重 + 纸箱 + 托盘）")

ctns = d.get("containers") or []
if len(ctns) > 1:  # 多柜：在第19行后插行并复制样式
    ws.insert_rows(20, amount=len(ctns) - 1)
    for i in range(1, len(ctns)):
        for col in range(1, 8):
            src = ws.cell(row=19, column=col)
            dst = ws.cell(row=19 + i, column=col)
            if src.has_style:
                dst._style = copy(src._style)

def fill_ctn(rn, c):
    ws.cell(row=rn, column=1).value = c.get("no", "")
    ws.cell(row=rn, column=2).value = c.get("seal", "")
    ws.cell(row=rn, column=3).value = c.get("type", "")
    ws.cell(row=rn, column=4).value = n(c.get("vgm"), 2)
    ws.cell(row=rn, column=5).value = n(c.get("pkgs"))
    ws.cell(row=rn, column=6).value = n(c.get("gw"), 2)
    ws.cell(row=rn, column=7).value = n(c.get("cbm"), 3)

for i, c in enumerate(ctns):
    fill_ctn(19 + i, c)

buf = io.BytesIO()
wb.save(buf)
sys.stdout.buffer.write(buf.getvalue())
